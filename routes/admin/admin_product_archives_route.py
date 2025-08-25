from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from datetime import datetime
from backend.database import DatabaseManager
import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image as PILImage

product_archives_bp = Blueprint('admin_product_archives', __name__, url_prefix='/admin/product_archives')

def login_required(f):
    def decorated_function(*args, **kwargs):
        from flask import session
        if 'admin_user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@product_archives_bp.route('/')
@login_required
def product_archives():
    """产品效果归档管理页面"""
    try:
        archives = DatabaseManager.get_product_archives()
    except Exception as e:
        print(f"获取产品效果归档数据失败: {e}")
        archives = []
    return render_template('admin/product_archives.html', archives=archives)

@product_archives_bp.route('/add', methods=['POST'])
@login_required
def add_product_archive():
    """添加产品效果归档"""
    try:
        data = request.get_json()
        access_code = data.get('access_code', '')
        original_product_image = data.get('original_product_image', '')
        original_depth_image = data.get('original_depth_image', '')
        effect_image = data.get('effect_image', '')
        effect_category = data.get('effect_category', '')
        registration_info = data.get('registration_info', '')
        follow_up_person = data.get('follow_up_person', '')
        
        if not all([access_code, original_product_image, original_depth_image, effect_image, effect_category]):
            return jsonify({'success': False, 'message': '请填写完整的归档信息'})
        
        # 这里应该处理文件上传，暂时使用占位符路径
        original_product_path = f"uploads/archives/original_product_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        original_depth_path = f"uploads/archives/original_depth_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        effect_image_path = f"uploads/archives/effect_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        archive_id = DatabaseManager.add_product_archive(
            access_code, original_product_image, original_depth_image, effect_image,
            effect_category, registration_info, follow_up_person,
            original_product_path, original_depth_path, effect_image_path
        )
        
        return jsonify({
            'success': True,
            'message': '产品效果归档添加成功！',
            'archive_id': archive_id
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})

@product_archives_bp.route('/get')
@login_required
def get_product_archive():
    """获取单个产品效果归档信息"""
    try:
        archive_id = request.args.get('id', type=int)
        if not archive_id:
            return jsonify({'success': False, 'message': '缺少归档ID'})
        
        archive = DatabaseManager.get_product_archive_by_id(archive_id)
        
        if archive:
            return jsonify({'success': True, 'data': archive})
        else:
            return jsonify({'success': False, 'message': '归档不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

@product_archives_bp.route('/delete', methods=['POST'])
@login_required
def delete_product_archive():
    """删除产品效果归档"""
    try:
        data = request.get_json()
        archive_id = data.get('id')
        
        if not archive_id:
            return jsonify({'success': False, 'message': '缺少归档ID'})
        
        # 先获取归档信息，以便删除对应的图片文件
        archive = DatabaseManager.get_product_archive_by_id(archive_id)
        if not archive:
            return jsonify({'success': False, 'message': '归档不存在'})
        
        # 删除数据库记录
        result = DatabaseManager.delete_product_archive(archive_id)
        
        if result > 0:
            # 删除对应的图片文件
            image_files = [
                archive.get('original_product_path'),
                archive.get('original_depth_path'), 
                archive.get('effect_image_path')
            ]
            
            deleted_files = []
            failed_files = []
            
            for image_file in image_files:
                if image_file:
                    # 构建完整的文件路径
                    file_path = os.path.join('uploads', 'archives', image_file)
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                            deleted_files.append(image_file)
                        else:
                            # 文件不存在，记录但不报错
                            pass
                    except Exception as e:
                        failed_files.append(f"{image_file}: {str(e)}")
            
            # 构建返回消息
            message = '产品效果归档删除成功！'
            if deleted_files:
                message += f' 已删除 {len(deleted_files)} 个图片文件。'
            if failed_files:
                message += f' 但有 {len(failed_files)} 个图片文件删除失败。'
            
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': '归档不存在或已删除'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})

@product_archives_bp.route('/export', methods=['POST'])
@login_required
def export_excel():
    """导出产品效果归档为Excel，支持图片插入"""
    try:
        data = request.get_json()
        selected_fields = data.get('fields', [])
        
        if not selected_fields:
            return jsonify({'success': False, 'message': '请选择要导出的字段'})
        
        # 获取所有归档数据
        archives = DatabaseManager.get_product_archives()
        
        if not archives:
            return jsonify({'success': False, 'message': '没有数据可导出'})
        
        # 字段映射
        field_mapping = {
            'access_code': '授权码',
            'original_product_path': '原产品图',
            'original_depth_path': '原深度图', 
            'effect_image_path': '产品效果图',
            'effect_category': '效果类型',
            'register_time': '登记时间',
            'follow_up_person': '登记人',
            'register_info': '登记信息'
        }
        
        # 图片字段
        image_fields = ['original_product_path', 'original_depth_path', 'effect_image_path']
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "产品效果归档"
        
        # 写入表头
        headers = [field_mapping[field] for field in selected_fields if field in field_mapping]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 设置行高（为图片预留空间）
        # 设置样式
        from openpyxl.styles import Alignment, Font
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=12, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 设置数据行样式
        data_font = Font(name='微软雅黑', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row_idx, archive in enumerate(archives, 2):
            # 检查当前行是否包含图片字段
            has_image = any(field in image_fields for field in selected_fields)
            
            # 根据是否有图片设置不同的行高
            if has_image:
                ws.row_dimensions[row_idx].height = 60  # 图片行高度
            else:
                ws.row_dimensions[row_idx].height = 20  # 普通行高度
            
            for col_idx, field in enumerate(selected_fields, 1):
                if field not in field_mapping:
                    continue
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = data_alignment
                    
                if field in image_fields:
                    # 处理图片字段
                    image_filename = archive.get(field, '')
                    if image_filename:
                        # 数据库中存储的是文件名，需要构建完整路径
                        full_image_path = os.path.join('uploads', 'archives', image_filename)
                        if os.path.exists(full_image_path):
                            try:
                                # 直接插入原图片到Excel
                                excel_img = Image(full_image_path)
                                
                                # 调整图片大小，让图片占满单元格的大部分空间
                                excel_img.width = 80   # 减小图片宽度
                                excel_img.height = 50  # 减小图片高度
                                
                                # 计算单元格位置，让图片居中
                                excel_img.anchor = cell.coordinate
                                ws.add_image(excel_img)
                                        
                            except Exception as e:
                                # 如果图片处理失败，显示文件名
                                cell.value = image_filename
                        else:
                            cell.value = "图片不存在"
                    else:
                        cell.value = "无图片"
                else:
                    # 处理非图片字段
                    value = archive.get(field, '') or ''
                    # 处理时间格式
                    if field == 'register_time' and value:
                        value = str(value)[:16] if len(str(value)) > 16 else str(value)
                    cell.value = value
        
        # 调整列宽
        for col_idx, field in enumerate(selected_fields, 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            if field in image_fields:
                # 图片列设置固定宽度
                ws.column_dimensions[column_letter].width = 12
            else:
                # 非图片列根据内容调整宽度
                max_length = 0
                for row in range(1, len(archives) + 2):
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value:
                        # 中文字符按2个字符计算宽度
                        length = sum(2 if ord(char) > 127 else 1 for char in str(cell_value))
                        if length > max_length:
                            max_length = length
                
                # 设置合适的列宽，最小8，最大30
                adjusted_width = max(min(max_length + 2, 30), 8)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"产品效果归档_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'})

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS